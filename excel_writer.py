import pandas as pd
import os
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from datetime import datetime

EXCEL_FILE = "job_applications.xlsx"

COLUMNS = [
    "company",
    "job_title",
    "date_applied",
    "response_type",
    "subject",
    "email",
    "thread_id"
]

def save_to_excel(records):
    if os.path.exists(EXCEL_FILE):
        df_existing = pd.read_excel(EXCEL_FILE)
    else:
        df_existing = pd.DataFrame(columns=COLUMNS)

    if "thread_id" not in df_existing.columns:
        print("⚠️ Excel file missing 'thread_id' column. Recreating file.")
        df_existing = pd.DataFrame(columns=COLUMNS)

    new_records = []
    existing_ids = set(df_existing["thread_id"].astype(str).str.lower())

    for r in records:
        if r["thread_id"].lower() not in existing_ids:
            new_records.append(r)

    df_all = df_existing.copy()

    if new_records:
        df_new = pd.DataFrame(new_records)
        df_all = pd.concat([df_existing, df_new], ignore_index=True)
        df_all.to_excel(EXCEL_FILE, index=False)
        print(f"✔️ Saved {len(new_records)} new records to {EXCEL_FILE}")
    else:
        print("No new records to write.")

    # ✅ Always apply coloring to the current file


def apply_row_colors(df, ws):
    from openpyxl.styles import PatternFill
    from datetime import datetime

    if df.empty or "response_type" not in df.columns or "date_applied" not in df.columns:
        print("⚠️ Missing required columns.")
        return

    df["date_applied"] = pd.to_datetime(df["date_applied"], errors="coerce")
    today = datetime.utcnow()

    def tag(row):
        resp = str(row["response_type"]).strip().lower()
        if resp in {"accepted", "interview"}:
            return "positive"
        elif resp == "rejected":
            return "negative"
        elif "no" in resp:
            if pd.isnull(row["date_applied"]):
                return "waiting"
            days = (today - row["date_applied"]).days
            return "stale" if days > 10 else "waiting"
        return "unknown"

    tags = df.apply(tag, axis=1)

    color_map = {
        "positive": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),  # green
        "negative": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),  # red
        "stale":    PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid"),  # yellow
        "waiting":  PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid"),  # gray
    }

    for i, tag_value in enumerate(tags, start=2):  # skip header row
        fill = color_map.get(tag_value)
        if fill:
            for col in range(1, len(df.columns) + 1):
                ws.cell(row=i, column=col).fill = fill


import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import PatternFill
from datetime import datetime

EXCEL_FILE = "job_applications.xlsx"

def archive_old_no_response_entries():

    df = pd.read_excel(EXCEL_FILE)
    if "date_applied" not in df.columns or "response_type" not in df.columns:
        print("⚠️ Required columns missing.")
        return

    df["date_applied"] = pd.to_datetime(df["date_applied"], errors="coerce")
    today = datetime.utcnow()

    def should_archive(row):
        resp = str(row.get("response_type", "")).strip().lower()
        if "rejected" in resp:
            return True
        if "applied" in resp:
            return True
        if "no" in resp and not pd.isnull(row.get("date_applied")):
            return (today - row["date_applied"]).days > 10
        return False

    archived_df = df[df.apply(should_archive, axis=1)]
    active_df = df[~df.apply(should_archive, axis=1)]

    # Save active to Excel
    active_df.to_excel(EXCEL_FILE, index=False)

    # Load workbook again
    wb = load_workbook(EXCEL_FILE)
    ws_main = wb.active
    apply_row_colors(active_df, ws_main)

    # Write Archived Sheet
    if "Archived" in wb.sheetnames:
        del wb["Archived"]
    ws_archived = wb.create_sheet("Archived")

    for row in dataframe_to_rows(archived_df, index=False, header=True):
        ws_archived.append(row)

    apply_row_colors(archived_df, ws_archived)

    wb.save(EXCEL_FILE)
    print(f"📂 Archived {len(archived_df)} entries → 'Archived' sheet with colored rows.")
